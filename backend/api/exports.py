"""Tabular exports for risk reports and impact analysis (Phase 3).

Pure builders (data -> rows -> bytes) kept separate from the HTTP layer so they
unit-test without a server: CSV via the stdlib, XLSX via openpyxl. The route
layer in ``routes_export.py`` wraps these in FastAPI responses.
"""
from __future__ import annotations

import csv
import io
from typing import Any, Sequence

# Stable column order so exported files have a predictable schema.
RISK_HEADERS = ["type", "severity", "target", "file", "details"]
IMPACT_HEADERS = ["relation", "name", "file", "hops"]
SECURITY_HEADERS = ["rule_id", "severity", "file", "line", "message"]
REFACTOR_HEADERS = ["id", "type", "title", "severity", "effort", "target", "file", "rationale", "suggestion"]


def risks_to_rows(risks: Sequence[dict]) -> tuple[list[str], list[list[Any]]]:
    rows = [[r.get(h, "") for h in RISK_HEADERS] for r in risks]
    return RISK_HEADERS, rows


def security_to_rows(findings: Sequence[dict]) -> tuple[list[str], list[list[Any]]]:
    rows = [[f.get(h, "") for h in SECURITY_HEADERS] for f in findings]
    return SECURITY_HEADERS, rows


def refactor_to_rows(recs: Sequence[dict]) -> tuple[list[str], list[list[Any]]]:
    rows = [[r.get(h, "") for h in REFACTOR_HEADERS] for r in recs]
    return REFACTOR_HEADERS, rows


def impact_to_rows(impact: dict) -> tuple[list[str], list[list[Any]]]:
    """Flatten the impact summary's direct/transitive entity lists into rows."""
    rows: list[list[Any]] = []
    for relation, key in (("direct", "directly_affected"), ("transitive", "transitively_affected")):
        for entity in impact.get(key, []):
            rows.append([relation, entity.get("name", ""), entity.get("file", ""), entity.get("hops", "")])
    return IMPACT_HEADERS, rows


def to_csv(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    # utf-8-sig so Excel opens non-ASCII content correctly on double-click.
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx(headers: Sequence[str], rows: Sequence[Sequence[Any]], sheet_title: str = "export") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "export"  # Excel caps sheet titles at 31 chars
    
    # Write headers
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    # Write rows with text wrap
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(list(row))
        for cell in ws[ws.max_row]:
            cell.alignment = wrap_alignment
            
    # Set column widths (approximate for typical content)
    col_widths = {"A": 15, "B": 10, "C": 25, "D": 40, "E": 60}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
        
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# MIME types + filename suffixes per format, used by the route layer.
FORMATS = {
    "csv": ("text/csv", "csv"),
    "xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
}


def render(fmt: str, headers: Sequence[str], rows: Sequence[Sequence[Any]],
           sheet_title: str = "export") -> bytes:
    if fmt == "csv":
        return to_csv(headers, rows)
    if fmt == "xlsx":
        return to_xlsx(headers, rows, sheet_title)
    raise ValueError(f"unsupported format: {fmt}")
